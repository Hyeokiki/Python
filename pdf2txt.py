from fileinput import filename
from io import StringIO

from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfdocument import PDFDocument
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.pdfpage import PDFPage
from pdfminer.pdfparser import PDFParser
import csv
import pandas as pd
import numpy as np
from tqdm import tqdm 
from openpyxl.styles import Font
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill

def convert_pdf_to_string(file_path):
    output_string = StringIO()
    with open(file_path, 'rb') as in_file:
        parser = PDFParser(in_file)
        doc = PDFDocument(parser)
        rsrcmgr = PDFResourceManager()
        device = TextConverter(rsrcmgr, output_string, laparams= LAParams())
        interpreter = PDFPageInterpreter(rsrcmgr, device)
        for page in tqdm(PDFPage.create_pages(doc)):
            interpreter.process_page(page)
        
    return(output_string.getvalue())


if __name__ == '__main__':
    originMonth = int(input("출력할 달을 입력해주세요(ex. 2월이면 2를 입력.) : "))
    pdfName = "./(주)오데야 "+ str(originMonth) +"월사업소득자.pdf"
    
    text = convert_pdf_to_string(pdfName)
    text = text.replace('.', '')
    text = text.replace('\x0c','')
    table_of_contents_raw = text.split('\n')
    #print(table_of_contents_raw)
    #for txt in table_of_contents_raw :
    #    print(txt)
    #print(convert_pdf_to_string(pdfName))
    
    name_dict = {} #이름 dictionary (ex. 0 : 첫 번째 사람 이름 ...)
    people_dict = {} #각 사람 별 정보 dictionary
    dict_index = 0
                
    name_dict[dict_index] = '유진'
    people_dict['유진'] = []
    
    #940909 #왼쪽으로 이름 찾기 #'유진'님은 두 글자라 따로 넣어 줘야됨 ㅠ
    for txt in tqdm(table_of_contents_raw) :
        if "940909 " in txt :
            #print(txt[7:10])
            name_dict[dict_index] = txt[7:10]
            people_dict[txt[7:10]] = []
            dict_index += 1
            
    dict_index = 0
    remove_set = {''}
    for name in tqdm(people_dict):
        ujinFlag = True
        flag = False
        for txt in table_of_contents_raw :
            if txt == '유진' and ujinFlag == True:
                ujinFlag = False
                continue
            if txt.find("동,") != -1 :
                continue
            if txt.find("길") != -1 :
                continue
            #print(txt)
            if flag :
                people_dict[name].append(txt)                
            if txt == name:
                flag = True
            if flag == True and txt == '관리번호':
                flag = False
                dict_index += 1
        people_dict[name] = [p for p in people_dict[name] if p not in remove_set]
        people_dict[name] = [p for p in people_dict[name] if p.find(",") != -1]
        people_dict[name] = '/'.join(dict.fromkeys(people_dict[name]))
        #print(name)
        #print(people_dict[name])
    
    wb = load_workbook("./"+ str(originMonth) +"월 프리랜서 총합(완성본).xlsx")
    ws = wb.active
    
    #print(len(people_dict))
    maxRow = len(people_dict) + 4
    for rowIndex in tqdm(range(4, maxRow, 1)):
        targetName = ws.cell(row = rowIndex, column = 3).value
        if targetName in people_dict:
            #if targetName == '유진':
                #print(people_dict[targetName])
            if people_dict[targetName] != []:
                #print(people_dict[targetName].split('/')[-1])5
                ws.cell(row = rowIndex, column = 5).value = "₩" + people_dict[targetName].split('/')[-1]
    wb.save("./"+ str(originMonth) +"월 프리랜서 총합(완성본).xlsx")
    print("#######프리랜서 총합(완성본) 생성 완료#######")
   
    

    
