import pandas as pd
import numpy as np
import datetime
from tqdm import tqdm 
from openpyxl.styles import Font
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill
from openpyxl.chart import BarChart, Series, Reference
import sys


if __name__ == "__main__":
    #분당 돈 계산하고 소수점은 절삭
    #n월 n주차 구하기
    def week_no(y, m, d):
        #연월일을 입력받아 해당 요일의 주차를 얻는 함수
        '''y(int) - 연도 m(int) - 월 d(int) - 일, return : 해당 요일의 주(int)
        '''
        def _ymd_to_datetime(y, m, d): # 3
            """ 연월일을 입력받아 datetime 객체로 변환하는 함수
            Args:
            y (int) - 연도
            m (int) - 월
            d (int) - 일

            Return:
            datetime - YYYY-MM-DD 형식의 datetime 객체
            """
            s = f'{y:04d}-{m:02d}-{d:02d}'
            return datetime.datetime.strptime(s, '%Y-%m-%d')

        target_day = _ymd_to_datetime(y, m, d) # 4
        firstday = target_day.replace(day=1) # 5
        while firstday.weekday() != 0: # 6
            firstday += datetime.timedelta(days=1)
        if target_day < firstday: # 7
            return 0
    
        return (target_day - firstday).days // 7 + 1 # 8

    #print(str(week_no(2022, 3, 1))) test용 print()

    #요일 계산기
    def cal_day(y, m, d):
        days = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        #토 : 5 일 : 6
        b = days[datetime.date(y, m, d).weekday()]
        #b = datetime.date(y, m, d).weekday()
        return b
    #print(cal_day(2022, 3, 19))

    #급여 계산용 함수
    def cal_pay(name, ):
        payment = 0
        return payment

    originMonth = int(sys.argv[1]) #두 번째 인자 값에 월을 입력한다.
    #originMonth = 3
    originName = "2022년 "+ str(originMonth) +"월 출근시간표.xlsx"

    globalMonth = originMonth
    #df = pd.read_excel('22.02.worksheet.xlsx')
    df = pd.read_excel(originName)
    df = df.fillna('')
    #print(df['Unnamed: 2'][3:-1]) 날짜
    #print(df['Unnamed: 3'][3:-1]) 근무인원
    #print(df['Unnamed: 4'][3:-1]) 근무태그
    #workLoad = np.array(df['Unnamed: 5'][3:-1]) 근무량
    #print(df['Unnamed: 7'][3:-1]) 출근시간
    #print(df['Unnamed: 8'][3:-1]) 퇴근시간
    #print(df['Unnamed: 9'][3:-1]) 비고
    #print()
    #print(df)

    #print(df['Unnamed: 12'][3:-1])
    #print(df['Unnamed: 15'][3:-1])

    def add_comma(money):
        mSize = len(money)
        if mSize > 3 :
            check = 0
            for index in range(mSize, 0, -1) :
                if check%3 == 0  and check != 0 :
                    money = money[0 : index] + "," + money[index :]
                check += 1
        return money

    def check_end(y1, m1, day1 , y2, m2, day2):
        if int(m1) == int(m2):
            if week_no(y1, m1, day1) == week_no(y2, m2, day2):
                return True
            else:
                return False
        else : 
            return False

    def check_contract(l):
        result = 1
        for c in l:
            result = result * c
        if result == 0:
            return False
        else :
            return True

    Day = np.array(df['Unnamed: 1'][3:-1])
    workDay = np.array(df['Unnamed: 2'][3:-1])
    name = np.array(df['Unnamed: 3'][3:-1])
    workTag = np.array(df['Unnamed: 4'][3:-1])
    workLoad = np.array(df['Unnamed: 5'][3:-1])
    inRecord = np.array(df['Unnamed: 7'][3:-1])
    outRecord = np.array(df['Unnamed: 8'][3:-1])
    etc = np.array(df['Unnamed: 9'][3:-1]) #비고

    nameCheck = np.array(df['Unnamed: 12'][3:-1]) #수습 및 반장 확인용 리스트
    dateChange = np.array(df['Unnamed: 15'][3:-1]) #수습 본계약 전환일 리스트
    contractCheck = np.array(df['Unnamed: 16'][3:-1]) #주휴수당 여부 확인용 계약 이름 체크 리스트
    endCheck = np.array(df['Unnamed: 17'][3:-1]) #근무 종료일 체크

    tempList = []
    peopleList = []
    for index in range(0,len(name)):
        tempList.append(workDay[index])
        tempList.append(name[index])
        tempList.append(workTag[index])
        tempList.append(inRecord[index])
        tempList.append(outRecord[index])
        tempList.append(workLoad[index])
        tempList.append("") #식대 칸
        tempList.append(etc[index])
        #tempList.append(day[index]) #근무날의 요일을 추가
        peopleList.append(tempList)
        tempList = []


    commuteDf = pd.DataFrame(peopleList)
    #print(commuteDf)
    #commuteDf.to_excel("testResult.xlsx", index = False )

    #개인별 월별 급여내역서 만들기
    wage = [ 9160, 9500, 10000] #0 : 2022 최저 시급(수습), 1 : 현재 시급, 2 : 반장님 시급

    wage_dict = {} #개인별 급여 타입 dictionary
    wage_date_dict = {} #개인별 급여 변경 날짜 dictionary

    contract_dict = {} #개인별 근무 요일 dictionary
    contract_success_dict = {} #개인별 근무 요일 충족

    end_dict = {} #개인별 종료 dictionary

    record_dict = {} #이월되는 주휴수당 기록용 dict

    tempd = {}
    tempd['Mon'] = 0
    tempd['Tue'] = 1
    tempd['Wed'] = 2
    tempd['Thu'] = 3
    tempd['Fri'] = 4
    tempd['Sat'] = 5
    tempd['Sun'] = 6

    #계약 조건에서 요일 따지기
    
    for i in range(0, len(nameCheck)):
        if dateChange[i] == "M" :
            wage_date_dict[nameCheck[i]] = "M"
        elif nameCheck[i] != "" and nameCheck[i] != "성함" and dateChange[i] != "반장" and dateChange[i] != "":
            wage_date_dict[nameCheck[i]] = str(dateChange[i])[5:7] + "-" + str(dateChange[i])[8:10] #날짜 기록
            #print(nameCheck[i])
            #print(wage_date_dict[nameCheck[i]])
        
        if(endCheck[i] != ""):
            end_dict[nameCheck[i]] = str(endCheck[i])[5:7] + "-" + str(endCheck[i])[8:10] #종료 날짜 기록
    '''    
        if contractCheck[i] != "":
            contract_dict[nameCheck[i]] = contractCheck[i].split(',')
            #print(nameCheck[i])
            #print(contract_dict[nameCheck[i]])
            temp = [1,1,1,1,1,1,1]
            for d in contract_dict[nameCheck[i]]:
                temp[tempd[d]] = 0
            contract_success_dict[nameCheck[i]] = temp
    '''
        
            
    timeformat = "%H:%M" #시간 계산 포맷
    dateformat = "%m-%d" #날짜 계산 포맷

    line_dict_person = {} #사람마다 주마다 라인 줄 계산한 리스트 저장
    work_dict_person = {} #사람마다 주마다 성과 줄 계산한 리스트 저장
    for n in tqdm(name) :
        if n not in end_dict :
            end_dict[n] = "99-99"

        #if n not in contract_success_dict:
        #    contract_success_dict[n] = [1, 1, 1, 1, 1, 1, 1] #0은 계약 근무 날짜를 표현함
        wage_dict[n] = 9500 #일단 9500으로 초기화
        line_dict = {} #주차별 줄 수 계산
        line_dict[0] = 0 #이거 1주차 아님 젤 첫번째꺼 예외용
        line_dict[1] = 0
        line_dict[2] = 0
        line_dict[3] = 0
        line_dict[4] = 0
        line_dict[5] = 0
        work_dict_person[n] = []
        tempList = []
        personData = []
        sumWorkload = 0 #총 검수량
        sumWorkHour = 0 #총 근무 Hour
        sumWorkMinute = 0 #총 근무 Minute
        sumOverWorkHour = 0 #총 초과 근무 Hour
        sumOverWorkMinute = 0 #총 초과 근무 Minute
        sumWage = 0 #총 급여
        sumRiceCount = 0 #총 식대 횟수
        sumLastWage = 0 #총 이월 급여
        lastWorkHour_week = 0
        lastWorkMinute_week = 0
        lastOverWorkHour_week = 0
        lastOverWorkMinute_week = 0
        lastWage_week = 0 #마지막 주차 다음달 주휴수당을 제외한 급여
        
        lastRiceCount = 0 #총 이월 식대
        recordWage = 0 #이월되는 주휴수당 기록용
        
        tempList.append("성함 : ")
        tempList.append(n)
        personData.append(tempList)
        personData.append([])
        
        tempList = []
        tempList.append("날짜")
        tempList.append("근무파트")
        tempList.append("출근")
        tempList.append("퇴근")
        tempList.append("근무시간")
        tempList.append("검수평균")
        tempList.append("식대")
        tempList.append("비 고")
        personData.append(tempList)
        
        '''
        #주차별 근무 시간 기록용 dictionary
        workTimeWeek_dict = {}
        workTimeWeek_dict[0] = 0 #1주차 총 근무 시간
        workTimeWeek_dict[1] = 0 #2주차 총 근무 시간
        workTimeWeek_dict[2] = 0 #3주차 총 근무 시간
        workTimeWeek_dict[3] = 0 #4주차 총 근무 시간
        workTimeWeek_dict[4] = 0 #5주차 총 근무 시간
        
        #주차별 근무 성과 기록용 dictionary
        workLoadWeek_dict = {}
        workLoadWeek_dict[0] = 0 #1주차 총 근무 성과
        workLoadWeek_dict[1] = 0 #2주차 총 근무 성과
        workLoadWeek_dict[2] = 0 #3주차 총 근무 성과
        workLoadWeek_dict[3] = 0 #4주차 총 근무 성과
        workLoadWeek_dict[4] = 0 #5주차 총 근무 성과
        '''
        
        sumWorkHour_week = 0 #주차 별 근무 hour
        sumWorkMinute_week = 0 #주차 별 근무 minute
        sumOverWorkHour_week = 0 #주차 별 근무 hour
        sumOverWorkMinute_week = 0 #주차 별 근무 minute
        sumWorkload_week = 0 #주차 별 검수량
        sumWage_week = 0 #주차 별 급여
        sumRiceCount_week = 0 #주차 별 식대 횟수    
        
        dayCheck_week = True #주차 별 근무 요일 계약 조건 충족여부
        weekCheck = False #근무 처음 시작하는 주차 체크~
        weekFlag = True #근무 처음 시작하는 주차 출력용  
        HOLFlag = True #그 주에 Holiday만 있으면 주차 출력 x
        

        countWeek = 0 #주차 체크 변수
        for person in peopleList:
            if person[1] == n :
                tempWORKDAY = str(person[0])[5:7] + "-" + str(person[0])[8:10] #날짜 기록
                countWeek = week_no( 2022, int(tempWORKDAY[0:2]), int(tempWORKDAY[3:5]))
                break
        
        for person in peopleList :
            if person[1] == n :
                #주차 기록
                WORKDAY = str(person[0])[5:7] + "-" + str(person[0])[8:10] #날짜 기록
                lastCheck = True
                #print(int(WORKDAY.split("-")[0]))
                if globalMonth != int(WORKDAY.split("-")[0]):
                    lastCheck = False
                if weekFlag:
                    tempList = []
                    personData.append(tempList)
                    tempList = []
                    tempList.append(str(int(week_no( 2022,int(WORKDAY[0:2]),int(WORKDAY[3:5]))) + 1)+"주차")
                    personData.append(tempList)
                    weekFlag = False
                #주차가 바뀌는 파트
                if HOLFlag != True:
                    countWeek = week_no( 2022, int(WORKDAY[0:2]), int(WORKDAY[3:5]))
                elif countWeek != week_no( 2022,int(WORKDAY[0:2]),int(WORKDAY[3:5])) and weekCheck and lastCheck:
                    countWeek = week_no( 2022, int(WORKDAY[0:2]), int(WORKDAY[3:5]))
                    #weekFlag = False
                    if countWeek > 0 :
                        #주차 별 결산 기록
                        tempList = []
                        personData.append(tempList)
                        tempList = []
                        tempList.append(str(countWeek) + "주차 결산 : ")
                        tempList.append("")
                        tempList.append("")
                        tempList.append("")

                        if sumWorkMinute_week >= 60:
                            sumWorkHour_week += (int)(sumWorkMinute_week/60)
                            sumWorkMinute_week = (int)(sumWorkMinute_week%60)
                        
                        if sumOverWorkMinute_week >= 60:
                            sumOverWorkHour_week += (int)(sumOverWorkMinute_week/60)
                            sumOverWorkMinute_week = (int)(sumOverWorkMinute_week%60)
                        
                        changeHour = (int)((sumWorkMinute_week+sumOverWorkMinute_week)/60)
                        changeMinute = (sumWorkMinute_week+sumOverWorkMinute_week)%60
                        tempList.append(str(sumWorkHour_week+sumOverWorkHour_week + changeHour) + "시간 " + str(changeMinute) + "분")
                        

                        #workTimeWeek_dict[countWeek] = sumWorktime_week
                        tempList.append(sumWorkload_week)
                        personData.append(tempList)
                        
                        #해당 주차 주휴수당 여부
                        weekBool1 = False #주 15시간 이상 주 40시간 미만 해당 여부
                        weekBool2 = False #주 40시간 이상 해당 여부
                        if (int(sumWorkHour_week) + int(sumOverWorkHour_week)) >= 15 and (int(sumWorkHour_week) + int(sumOverWorkHour_week)) < 40 :
                            weekBool1 = True
                        elif (int(sumWorkHour_week)+int(sumOverWorkHour_week)) >= 40 :
                            weekBool2 = True
                        
                        if not dayCheck_week : 
                            weekBool1 = False
                            weekBool2 = False
                        
                        #print(person[1])
                        #print(tempList)
                        #print(contract_success_dict[person[1]])
                        #print(check_contract(contract_success_dict[person[1]]))
                        '''
                        if not check_contract(contract_success_dict[person[1]]):
                            weekBool1 = False
                            weekBool2 = False
                            '''
                        if check_end(2022,int(WORKDAY[0:2]),int(WORKDAY[3:5]),2022,int(end_dict[n][0:2]),int(end_dict[n][3:5])):
                            weekBool1 = False
                            weekBool2 = False
                            
                        #사람 타입 별 계산
                        MONTH = int(WORKDAY.split("-")[0])
                        DAY = int(WORKDAY.split("-")[1])
                        #밑의 날짜는 본 계약 시작일
                        if wage_date_dict[n] != "M":
                            MONTH2 = int(wage_date_dict[n].split("-")[0])
                            DAY2 = int(wage_date_dict[n].split("-")[1])
                            if MONTH < MONTH2 :
                                wage_dict[n] = wage[0] #수습
                                weekBool1 = False
                                weekBool2 = False
                            elif MONTH == MONTH2 and DAY < DAY2 :
                                weekBool1 = False
                                weekBool2 = False

                        tempList = []
                        if weekBool1 == False and weekBool2 == False:
                            tempList.append("주휴수당 여부(주 15시간 이상 40시간 미만) : ")
                            tempList.append("")
                            tempList.append("")
                            tempList.append("")
                            tempList.append("")
                            tempList.append("❌")
                            personData.append(tempList)
                            tempList = []
                            tempList.append("주휴수당 여부(주 40시간 이상) : ")
                            tempList.append("")
                            tempList.append("")
                            tempList.append("")
                            tempList.append("")
                            tempList.append("❌") 
                        if weekBool1 :
                            tempList.append("주휴수당 여부(주 15시간 이상 40시간 미만) : ")
                            tempList.append("")
                            tempList.append("")
                            tempList.append("")
                            tempList.append("")
                            tempList.append("⭕️")
                            personData.append(tempList)
                            tempList = []
                            tempList.append("주휴수당 여부(주 40시간 이상) : ")
                            tempList.append("")
                            tempList.append("")
                            tempList.append("")
                            tempList.append("")
                            tempList.append("❌") 
                        if weekBool2 : 
                            tempList.append("주휴수당 여부(주 40시간 이상) : ")
                            tempList.append("")
                            tempList.append("")
                            tempList.append("")
                            tempList.append("")
                            tempList.append("❌")
                            personData.append(tempList)
                            tempList = []
                            tempList.append("주휴수당 여부(주 15시간 이상 40시간 미만) : ")
                            tempList.append("")
                            tempList.append("")
                            tempList.append("")
                            tempList.append("")
                            tempList.append("⭕️")
                        personData.append(tempList)
                    
                        #해당 주차 급여 계산
                        
                        #### FLAG ####                                    
                        tempWage_week = 0 #한 주의 주휴수당 기록용
                        tempList = []
                        tempList.append(str(countWeek) + "주차 급여(주휴수당 제외) : ")
                        if weekBool1 == True : #주휴수당 15시간 이상 여부
                            tempWage_week = (float)( (( (sumWorkHour_week + (float)(sumWorkMinute_week )/60) / 40)) * 8 * personWage )
                            tempWage_week += (float)(( (sumOverWorkHour_week + (float)(sumOverWorkMinute_week / 60))/40) * 8 * ((float)(personWage)))
                        elif weekBool2 == True : #주휴수당 40시간 이상 여부
                            tempWage_week = 8 * personWage
                            #tempWage_week += 8 * ((float)(personWage * 1.5))
                        tempList.append("")
                        tempList.append("")
                        tempList.append("")
                        tempList.append("")
                        #for i in range(len(sumWage_week)-1, -1, -1):
                        #    if (i % 3) == 2:
                        #    sumWage_week[i]
                        tempList.append("₩" + add_comma(str(int(sumWage_week))))
                        personData.append(tempList)
                        tempList = []
                        
                        tempList.append(str(countWeek) + "주차 주휴 수당 : ")
                        tempList.append("")
                        tempList.append("")
                        tempList.append("")
                        tempList.append("")
                        tempList.append("₩" + add_comma(str(int(tempWage_week))))
                        personData.append(tempList)
                        
                        tempList = []
                        tempList.append(str(countWeek) + "주차 식대 : ")
                        tempList.append("")
                        tempList.append("")
                        tempList.append("")
                        tempList.append("")
                        tempList.append("₩" + add_comma(str(int(sumRiceCount_week) * 9000)))
                        personData.append(tempList)
                    
                        tempList = []
                        sumWage_week += tempWage_week
                        sumWage_week += int(sumRiceCount_week) * 9000
                        sumWage += sumWage_week
                        sumWorkHour_week = 0 #주차별 근무 시간 초기화
                        sumOverWorkHour_week = 0 #주차별 초과 근무 시간 초기화
                        sumWorkload_week = 0 #주차별 근무 성과 
                        sumWorkMinute_week = 0
                        sumOverWorkMinute_week = 0
                        sumRiceCount_week = 0
                        dayCheck_week = True
                        
                        '''
                        #근무 요일 체크 초기화
                        temp = [1,1,1,1,1,1,1]
                        if person[1] in contract_dict:
                            for d in contract_dict[person[1]]:
                                temp[tempd[d]] = 0
                            contract_success_dict[person[1]] = temp
                        '''
                        tempList.append(str(countWeek) + "주차 총 급여 =================> ")
                        tempList.append("")
                        tempList.append("")
                        tempList.append("")
                        tempList.append("")
                        tempList.append("₩" + add_comma(str(int(sumWage_week))))
                        personData.append(tempList)
                                        
                    #다음 주차 마크 기록
                    #if not weekFlag :
                    tempList = []
                    personData.append(tempList)
                    tempList = []
                    tempList.append(str(countWeek + 1)+"주차")
                    personData.append(tempList)
                    sumWage_week = 0
                    sumRiceCount_week = 0
                #근무 요일 계약 조건 충족 여부 판단
                if person[1] in nameCheck :
                    dayIndex = cal_day(2022, int(WORKDAY[0:2]),int(WORKDAY[3:5])) 
                    #contract_success_dict[person[1]][tempd[dayIndex]] = 1
                    #if cal_day(2022, int(WORKDAY[0:2]),int(WORKDAY[3:5])) not in cont ract_dict[person[1]] :
                    #   dayCheck_week = False
                if person[7] == "HOL":
                    dayIndex = cal_day(2022, int(WORKDAY[0:2]),int(WORKDAY[3:5])) 
                    #print(dayIndex)
                    #contract_success_dict[person[1]][tempd[dayIndex]] = 1
                    HOLFlag = False
                    continue
                else :
                    HOLFlag = True
                #날짜 기록
                tempList = []
                tempList.append(WORKDAY + " " + str(cal_day(2022, int(WORKDAY[0:2]),int(WORKDAY[3:5]))))
                
                #첫 근무 시작하는 주 체크
                if HOLFlag == True:
                    weekCheck = True
                
                #출근 시간 기록
                tempList.append(person[2]) #근무 타입 태그
                
                INTIME = str(person[3])[0:5]
                tempList.append(INTIME)
                
                #퇴근 시간 기록
                OUTTIME = str(person[4])[0:5]
                tempList.append(OUTTIME)
                
                #초과 근무 체크
                
                #근무 시간 계산
                intime_convert = datetime.datetime.strptime(INTIME, timeformat)
                outtime_convert = datetime.datetime.strptime(OUTTIME, timeformat)
                cal_workHour = (outtime_convert - intime_convert).seconds / 3600 #시 계산
                cal_workHour = (int)(cal_workHour)
                cal_workMinute = (outtime_convert - intime_convert).seconds / 60 #분 계산
                cal_workMinute = (int)(cal_workMinute%60)
                
                #풀타임 근무자 점심시간 1시간 제외
                if person[2] == "풀":
                    cal_workHour -= 1
                
                
                #근무 시간 합
                tempList.append(str(cal_workHour)+"시간 " + str(cal_workMinute) + "분")
                
                #사람 타입 별 계산
                if n in nameCheck :
                    MONTH = int(WORKDAY.split("-")[0])
                    DAY = int(WORKDAY.split("-")[1])
                    #밑의 날짜는 본 계약 시작일
                    if wage_date_dict[n] != "M":
                        MONTH2 = int(wage_date_dict[n].split("-")[0])
                        DAY2 = int(wage_date_dict[n].split("-")[1])
                        #print(n)
                        #print(str(MONTH) + " - "+ str(DAY))
                        #print(str(MONTH2) + " - "+ str(DAY2))

                    if wage_date_dict[n] == "M":
                        wage_dict[n] = wage[2] #반장님
                    elif MONTH > MONTH2 :
                        wage_dict[n] = wage[1] #9500원
                    elif MONTH < MONTH2 :
                        wage_dict[n] = wage[0] #수습
                    elif DAY < DAY2 :
                        wage_dict[n] = wage[0] #수습
                    else :
                        wage_dict[n] = wage[1] #9500원
                    #print(wage_dict[n])
                    
                personWage = wage_dict[n]
                
                #다음달인 근무 시간 합 구하기
                if globalMonth != int(WORKDAY.split("-")[0]) and person[7] != "DO":
                    lastWorkHour_week += cal_workHour
                    lastWorkMinute_week += cal_workMinute
                    
                    sumWorkHour -= cal_workHour
                    sumWorkHour_week -= cal_workHour
                    sumWorkMinute -= cal_workMinute 
                    sumWorkMinute_week -= cal_workMinute
                    
                    lastWage_week += personWage * cal_workHour
                    lastWage_week += ((float)(cal_workMinute/60)) * personWage
                    
                elif globalMonth != int(WORKDAY.split("-")[0]) and person[7] == "DO":
                    lastOverWorkHour_week += cal_workHour
                    lastOverWorkMinute_week += cal_workMinute
                    
                    sumOverWorkHour -= cal_workHour
                    sumOverWorkHour_week -= cal_workHour
                    sumOverWorkMinute -= cal_workMinute 
                    sumOverWorkMinute_week -= cal_workMinute 
                    
                    lastWage_week += int(personWage * 1.5) * cal_workHour
                    lastWage_week += ((float)(cal_workMinute/60)) * (personWage * 1.5)
                elif person[7] == "HOUSE":
                    sumWage_week += 9500 * cal_workHour
                    sumWage_week += float(cal_workMinute/60) * 9500    
                elif person[7] != "DO" : 
                    sumWage_week += personWage * cal_workHour
                    sumWage_week += float(cal_workMinute/60) * personWage
                elif person[7] == "DO" :
                    sumWage_week += float(personWage * 1.5) * cal_workHour
                    sumWage_week += float(cal_workMinute/60) * float(personWage * 1.5)
                
                #if person[1] == "조수미" and person[7] == "HOUSE":
                #    print()
                #   print(WORKDAY)
                #   print(9500 * cal_workHour + float(cal_workMinute/60) * 9500)
                #if person[1] == "조수미":
                #    print()
                #    print(WORKDAY)
                #    print(9500 * cal_workHour + float(cal_workMinute/60) * 9500)
                
                if person[7] != "DO":
                    sumWorkHour += cal_workHour
                    sumWorkHour_week += cal_workHour
                    sumWorkMinute += cal_workMinute 
                    sumWorkMinute_week += cal_workMinute #주마다 분 더하기
                elif person[7] == "DO" :
                    sumOverWorkHour += cal_workHour
                    sumOverWorkHour_week += cal_workHour
                    sumOverWorkMinute += cal_workMinute 
                    sumOverWorkMinute_week += cal_workMinute #주마다 분 더하기
                #근무 주차 계산
                #workWeek_dict[countWeek] += cal_worktime
                
                #총 검수 평균 계산
                WORKLOAD = str(person[5]).split(".")[0]
                if WORKLOAD == "알 수 없음" or WORKLOAD == "해당x":
                    sumWorkload += 0
                    sumWorkload_week += 0
                    tempList.append(0)
                    work_dict_person[n].append([WORKDAY,0])
                else : 
                    sumWorkload += int(WORKLOAD)
                    sumWorkload_week += int(WORKLOAD)
                    tempList.append(int(WORKLOAD))
                    work_dict_person[n].append([WORKDAY,int(WORKLOAD)])
                #식대 여부 계산
                if person[2] != "풀" :
                    tempList.append("X") #식대 추가
                else :
                    #사람 타입 별 계산
                    flag = True
                    if person[1] in nameCheck :
                        MONTH = int(WORKDAY.split("-")[0])
                        DAY = int(WORKDAY.split("-")[1])
                        #밑의 날짜는 본 계약 시작일
                        if wage_date_dict[person[1]] != "M":
                            MONTH2 = int(wage_date_dict[person[1]].split("-")[0])
                            DAY2 = int(wage_date_dict[person[1]].split("-")[1])
                        if wage_date_dict[person[1]] == "M":
                            flag = True #반장님
                        elif MONTH > MONTH2 :
                            flag = True #수습아님
                        elif MONTH < MONTH2 :
                            flag = False #수습
                        elif DAY < DAY2 :
                            flag = False #수습
                        else :
                            flag = True #수습아님
                    if flag :
                        tempList.append("O") #식대 지급 o
                        
                        if globalMonth != int(WORKDAY.split("-")[0]) :
                            lastRiceCount += 1
                        else :
                            sumRiceCount_week += 1 #식대 횟수 카운트
                            sumRiceCount += 1
                    else :
                        tempList.append("X") #식대 지급 x
                
                #비고추가
                tempList.append(person[7]) #비고 추가
                personData.append(tempList)
                #print(person[1])
                #print(tempList)
                line_dict[countWeek+1] += 1 
                
                
        #마지막 주차 결산 기록
        countWeek = week_no( 2022, int(WORKDAY[0:2]), int(WORKDAY[3:5]))
        if countWeek >= 0 :
            if not lastCheck :
                countWeek += 4
            tempList = []
            personData.append(tempList)
            tempList = []
            tempList.append(str(countWeek+1) + "주차 결산 : ")
            tempList.append("")
            tempList.append("")
            tempList.append("")
            if sumWorkMinute_week >= 60:
                sumWorkHour_week += (int)(sumWorkMinute_week/60)
                sumWorkMinute_week = (int)(sumWorkMinute_week%60)
            if sumOverWorkMinute_week >= 60:
                sumOverWorkHour_week += (int)(sumOverWorkMinute_week/60)
                sumOverWorkMinute_week = (int)(sumOverWorkMinute_week%60)    
            changeHour = (int)((sumWorkMinute_week+sumOverWorkMinute_week)/60)
            changeMinute = (sumWorkMinute_week+sumOverWorkMinute_week)%60
            tempList.append(str(sumWorkHour_week+sumOverWorkHour_week + changeHour) + "시간 " + str(changeMinute) + "분")
            #tempList.append(str(sumWorkHour_week+sumOverWorkHour_week) + "시간 " + str(sumWorkMinute_week+sumOverWorkMinute_week) + "분")
            tempList.append(sumWorkload_week)
            personData.append(tempList)
            
            #해당 주차 주휴수당 여부
            weekBool1 = False #주 15시간 이상 주 40시간 미만 해당 여부
            weekBool2 = False #주 40시간 이상 해당 여부
            if int(sumWorkHour_week+lastWorkHour_week) >= 15 and int(sumWorkHour_week+lastWorkHour_week) < 40 :
                weekBool1 = True
            elif int(sumWorkHour_week+lastWorkHour_week) >= 40 :
                weekBool2 = True
                
            '''
            #근무 요일 계약 조건 충족 여부 판단
            if n in nameCheck :
                #print(n)
                #print(cal_day(2022, int(WORKDAY[0:2]),int(WORKDAY[3:5])))
                #print(contract_dict[n] )
                if cal_day(2022, int(WORKDAY[0:2]),int(WORKDAY[3:5])) not in contract_dict[n] :
                    #print("있음")
                    dayCheck_week = False
            '''
            if not dayCheck_week : 
                weekBool1 = False
                weekBool2 = False
            '''
            if not check_contract(contract_success_dict[person[1]]):
                weekBool1 = False
                weekBool2 = False
            '''
            if check_end(2022,int(WORKDAY[0:2]),int(WORKDAY[3:5]),2022,int(end_dict[n][0:2]),int(end_dict[n][3:5])):
                weekBool1 = False
                weekBool2 = False
                
            #사람 타입 별 계산
            MONTH = int(WORKDAY.split("-")[0])
            DAY = int(WORKDAY.split("-")[1])
            #밑의 날짜는 본 계약 시작일
            if wage_date_dict[n] != "M":
                MONTH2 = int(wage_date_dict[n].split("-")[0])
                DAY2 = int(wage_date_dict[n].split("-")[1])
                if MONTH < MONTH2 :
                    wage_dict[n] = wage[0] #수습
                    weekBool1 = False
                    weekBool2 = False
                elif MONTH == MONTH2 and DAY < DAY2 :
                    weekBool1 = False
                    weekBool2 = False
            tempList = []
            if weekBool1 == False and weekBool2 == False:
                tempList.append("주휴수당 여부(주 15시간 이상 40시간 미만) : ")
                tempList.append("")
                tempList.append("")
                tempList.append("")
                tempList.append("")
                tempList.append("❌")
                personData.append(tempList)
                tempList = []
                tempList.append("주휴수당 여부(주 40시간 이상) : ")
                tempList.append("")
                tempList.append("")
                tempList.append("")
                tempList.append("")
                tempList.append("❌")
                personData.append(tempList)
            if weekBool1 :
                tempList.append("주휴수당 여부(주 15시간 이상 40시간 미만) : ")
                tempList.append("")
                tempList.append("")
                tempList.append("")
                tempList.append("")
                tempList.append("⭕️")
                personData.append(tempList)
                tempList = []
                tempList.append("주휴수당 여부(주 40시간 이상) : ")
                tempList.append("")
                tempList.append("")
                tempList.append("")
                tempList.append("")
                tempList.append("❌") 
                personData.append(tempList)
            if weekBool2 : 
                tempList.append("주휴수당 여부(주 40시간 이상) : ")
                tempList.append("")
                tempList.append("")
                tempList.append("")
                tempList.append("")
                tempList.append("❌")
                personData.append(tempList)
                tempList = []
                tempList.append("주휴수당 여부(주 15시간 이상 40시간 미만) : ")
                tempList.append("")
                tempList.append("")
                tempList.append("")
                tempList.append("")
                tempList.append("⭕️")
                personData.append(tempList)
            
            #해당 주차 급여 계산
            #사람 타입 별 계산
            if n in nameCheck :
                MONTH = int(WORKDAY.split("-")[0])
                DAY = int(WORKDAY.split("-")[1])
                #밑의 날짜는 본 계약 시작일
                if wage_date_dict[n] != "M":
                    MONTH2 = int(wage_date_dict[n].split("-")[0])
                    DAY2 = int(wage_date_dict[n].split("-")[1])
                if wage_date_dict[n] == "M":
                    wage_dict[n] = wage[2] #반장님
                elif MONTH > MONTH2 :
                    wage_dict[n] = wage[1]
                elif MONTH < MONTH2 :
                    wage_dict[n] = wage[0] #수습
                elif DAY < DAY2 :
                    wage_dict[n] = wage[0] #수습
                else :
                    wage_dict[n] = wage[1]

            personWage = wage_dict[n]
            
            #sumWage_week = personWage * sumWorkHour_week
            #sumWage_week += ((int)(personWage*1.5)) * sumOverWorkHour_week
            #sumWage_week += personWage * (float)(sumWorkMinute_week/60)
            #sumWage_week += (float)(1.5 * personWage) * (float)(sumOverWorkMinute_week/60) 
            
            #lastWage_week = personWage * lastWorkHour_week
            #lastWage_week += float(personWage * 1.5) * lastOverWorkHour_week
            #lastWage_week += float(lastWorkMinute_week/60) * personWage
            #lastWage_week += float(lastOverWorkMinute_week/60) * float(personWage * 1.5)
            sumLastWage += lastWage_week
            
            tempWage_week = 0 #한 주의 주휴수당 기록용
            tempList = []
            tempList.append(str(countWeek+1) + "주차 급여(주휴수당 제외) : ")
            
            if globalMonth != int(WORKDAY.split("-")[0]):
                sumWorkHour_week += lastWorkHour_week
                sumWorkMinute_week += lastWorkMinute_week
                sumOverWorkHour_week += lastOverWorkHour_week
                sumOverWorkMinute_week += lastOverWorkMinute_week
                
            if weekBool1 == True : #주휴수당 15시간 이상 여부
                tempWage_week = ( (sumWorkHour_week + (float)(sumWorkMinute_week/60) )/40 * 8 * personWage)
                tempWage_week += ((sumOverWorkHour_week + (float)(sumOverWorkMinute_week/60))/40 * 8 * ((float)(personWage)))
            elif weekBool2 == True : #주휴수당 40시간 이상 여부
                tempWage_week = 8 * personWage
                #tempWage_week += 8 * ((float)(personWage*1.5))

            if globalMonth != int(WORKDAY.split("-")[0]):
                sumWorkHour_week -= lastWorkHour_week
                sumWorkMinute_week -= lastWorkMinute_week
                sumOverWorkHour_week -= lastOverWorkHour_week
                sumOverWorkMinute_week -= lastOverWorkMinute_week
            
            tempList.append("")
            tempList.append("")
            tempList.append("")
            tempList.append("")
            #for i in range(len(sumWage_week)-1, -1, -1):
            #    if (i % 3) == 2:
            #    sumWage_week[i]
            tempList.append("₩" + add_comma(str(int(sumWage_week))))
            personData.append(tempList)
            tempList = []

            tempList.append(str(countWeek+1) + "주차 주휴 수당 : ")
            tempList.append("")
            tempList.append("")
            tempList.append("")
            tempList.append("")
            if lastCheck :
                tempList.append("₩" + add_comma(str(int(tempWage_week))))
                recordWage = 0
                record_dict[n] = recordWage
            else :
                tempList.append("₩" + add_comma(str(int(tempWage_week))) + "(*********다음달 지급******)")
                recordWage = tempWage_week
                record_dict[n] = recordWage
                tempWage_week = 0
            personData.append(tempList)

            tempList = []
            tempList.append(str(countWeek+1) + "주차 식대 : ")
            tempList.append("")
            tempList.append("")
            tempList.append("")
            tempList.append("")
            tempList.append("₩" + add_comma(str(int(sumRiceCount_week) * 9000)))
            personData.append(tempList)
            
            tempList = []
            tempList.append("다음 " + str(globalMonth+1) + "월 이월 급여 : ")
            tempList.append("")
            tempList.append("")
            tempList.append("")
            tempList.append("")
            tempList.append("₩" + add_comma(str(int(sumLastWage))) + "(*********다음달 지급******)")
            line_dict[countWeek+1] += 1
            personData.append(tempList)

            tempList = []
            tempList.append("다음 " + str(globalMonth+1) + "월 이월 식대 : ")
            tempList.append("")
            tempList.append("")
            tempList.append("")
            tempList.append("")
            tempList.append("₩" + add_comma(str(int(lastRiceCount) * 9000)) + "(*********다음달 지급******)")
            line_dict[countWeek+1] += 1
            personData.append(tempList)
                
            tempList = []
            sumWage_week += tempWage_week
            sumWage_week += int(sumRiceCount_week) * 9000
            sumWage += sumWage_week
            sumWorkHour_week = 0 #주차별 근무 시간 초기화
            sumOverWorkHour_week = 0
            sumWorkload_week = 0 #주차별 근무 성과 
            sumWorkMinute_week = 0
            sumOverWorkMinute_week = 0
            sumRiceCount_week = 0
            
            tempList.append(str(countWeek+1) + "주차 총 급여 =================> ")
            tempList.append("")
            tempList.append("")
            tempList.append("")
            tempList.append("")
            if not lastCheck :
                sumWage_week -= tempWage_week
                #tempList.append("₩" + add_comma(str(int(sumWage_week-tempWage_week))))
            tempList.append("₩" + add_comma(str(int(sumWage_week))))
            personData.append(tempList)
            
            sumWorkHour_week = 0
            sumWorkMinute_week = 0
            sumOverWorkHour_week = 0
            sumOverWorkMinute_week = 0
            sumWorkload_week = 0    
        #print(n)
        #print(line_dict[1])
        #print(line_dict[2])
        #print(line_dict[3])
        #print(line_dict[4])
        #print(line_dict[5])
        line_dict_person[n] = [line_dict[1]+line_dict[0], line_dict[2], line_dict[3], line_dict[4], line_dict[5]]
        personData.append([])
        personData.append(["**********************************************************************************"])
        tempList = []
        tempList.append("해당 월 총 성과 : ") 
        tempList.append("")
        tempList.append("")
        if sumWorkMinute >= 60:
            sumWorkHour += (int)(sumWorkMinute/60)
            sumWorkMinute = (int)(sumWorkMinute%60)
        if sumOverWorkMinute >= 60:
            sumOverWorkHour += (int)(sumOverWorkMinute/60)
            sumOverWorkMinute = (int)(sumOverWorkMinute%60)
        
        if sumWorkHour+sumOverWorkHour != 0:
            tempList.append(sumWorkload/(sumWorkHour+sumOverWorkHour)) #성과 평균 계산
        else :
            tempList.append(0)
        #tempList.append(sumWorkHour)
        tempList.append(str(sumWorkHour + sumOverWorkHour) + "시간 " + str(sumWorkMinute+sumOverWorkMinute) + "분")
        tempList.append(sumWorkload)
        personData.append(tempList)
        
        lastMonthWage = 0
        try :
            wb = load_workbook("./"+ str(originMonth-1) +"월/" + n + "_" + str(originMonth-1) + "월 급여내역서.xlsx")
            ws = wb.active
        except FileNotFoundError :
            lastMonthWage = 0
        else :
            lastMonthWage = int(ws.cell(row = 2, column = 4).value)
        
        personData.append([])
        tempList = []
        tempList.append(str(globalMonth-1) + "월 이월 주휴수당 : ")
        tempList.append("")
        tempList.append("")
        tempList.append("₩" + add_comma(str(lastMonthWage)))
        sumWage += lastMonthWage
        personData.append(tempList)
        
        personData.append([])
        tempList = []
        tempList.append(str(globalMonth) + "월 총 급여 : ")
        tempList.append("")
        tempList.append("")
        tempList.append("₩" + add_comma(str(int(sumWage))))
        personData.append(tempList)
        
        personData.append([])
        tempList = []
        tempList.append("3.3% 공제액 : ")
        tempList.append("")
        tempList.append("")
        TAX = (int)(sumWage * 0.033)
        tempList.append("₩" + add_comma(str(int(TAX))))
        personData.append(tempList)
        
        personData.append([])
        tempList = []
        tempList.append("프리랜서 신고 후 금액 : ")
        tempList.append("")
        tempList.append("")
        tempList.append("₩" + add_comma(str(int(sumWage - TAX))))
        personData.append(tempList)
        personData.append(["**********************************************************************************"])
        
        personalDf = pd.DataFrame(personData)
        personalDf.to_excel("./"+ str(originMonth) +"월/" + n + "_" + str(originMonth) + "월 급여내역서.xlsx", index = False)
    '''           
    for i in work_dict_person :
        print(i)
        print(work_dict_person[i])
    '''

# 이 파트는 엑셀 수정 및 출력 파트임.

for n in tqdm(name) :
    wb = load_workbook("./"+ str(originMonth) +"월/" + n + "_" + str(originMonth) + "월 급여내역서.xlsx")
    ws = wb.active
    
    #이월되는 주휴수당 기록
    ws.cell(row = 2, column = 4).value = int(record_dict[n])
    ws.cell(row = 2, column = 4).font = Font(color = "FFFFFF")
    #검수 평균 차트 그리기
    ws['L4'] = '날짜'
    ws['M4'] = '검수평균'
    index = 0
    for i in range(5, len(work_dict_person[n])+5, 1):
        ws.cell(row = i, column = 12).value = work_dict_person[n][index][0]
        ws.cell(row = i, column = 13).value = int(work_dict_person[n][index][1])
        #print(n)
        #print(work_dict_person[n][index][0])
        #print(work_dict_person[n][index][1])
        index += 1
    wb.save("./"+ str(originMonth) +"월/" + n + "_" + str(originMonth) + "월 급여내역서.xlsx")
    
    wb = load_workbook("./"+ str(originMonth) +"월/" + n + "_" + str(originMonth) + "월 급여내역서.xlsx")
    ws = wb.active
    c1 = BarChart()
    c1.type = "col"
    c1.style = 10
    c1.title = n + "님의 검수 평균"
    c1.y_axis.title = '검수 평균'
    c1.x_axis.title = '날 짜'
    data = Reference(ws, min_col = 13, min_row = 4, max_row = 4 + len(work_dict_person[n]))
    cats = Reference(ws, min_col=12, min_row=5, max_row= 5 + len(work_dict_person[n]))
    c1.add_data(data, titles_from_data=True)
    c1.set_categories(cats)
    c1.shape = 4
    ws.add_chart(c1, "O4")
    
    b2 = ws["B2"] #성함
    b2.font = Font(bold = True, size = 20)
    
    ws.column_dimensions['A'].width = 11 #근무날짜 열 너비 설정
    ws.column_dimensions['E'].width = 15 #근무시간 열 너비 설정
    
    #테두리 설정
    #max_row = ws.rows
    max_row = sum(line_dict_person[n]) + 56 #54는 근무 날짜 줄 수 빼고 남은 줄  수임
    if line_dict_person[n][4] != 0:
        max_row += 10
    for i in range(0, 4, 1):
        if line_dict_person[n][i] == 0 :
            max_row -= 10
    #print(n)
    #print(max_row)
    #print(sum(line_dict_person[n]))
    for j in range(1, 9, 1) :
        medium_border_top = Border(top = Side(style = 'medium'))
        ws.cell(row = 4, column = j).border = medium_border_top
        
    for i in range(5, max_row,1) :
        medium_border_right = Border(right = Side(style = 'medium'))
        ws.cell(row = i, column = 8).border = medium_border_right
        
    for i in range(5, max_row,1) :
        medium_border_left = Border(left = Side(style = 'medium'))
        ws.cell(row = i, column = 1).border = medium_border_left
        
    for j in range(1, 9, 1) :
        medium_border_bottom = Border(bottom = Side(style = 'medium'))
        ws.cell(row = max_row, column = j).border = medium_border_bottom
    
    h4 = ws["H4"]
    hm = ws["H"+str(max_row)]
    
    h4.border = Border(top = Side(style = 'medium'), right = Side(style = 'medium'))
    hm.border = Border(bottom = Side(style = 'medium'), right = Side(style = 'medium'))
    
    #날짜, 검수평균 ,근무시간 태그 서식
    for j in range(1, 9, 1):
        ws.cell(row = 4, column = j).fill = PatternFill(fgColor = "D5D5D5", fill_type = "solid")
        ws.cell(row = 4, column = j).font = Font(bold = True)
    
    line = 5
    if line_dict_person[n][0] != 0:
        line = 8+line_dict_person[n][0]+1+6
        ws.cell(row = 6, column = 1).value = "1주차"
        #주차 별 셀 색 채우기
        for i in range(6, line, 1):
            for j in range(1, 9 ,1):
                ws.cell(row = i, column = j).fill = PatternFill(fgColor = "EDFFE7", fill_type = "solid")
    if line_dict_person[n][1] != 0:
        line2 = line + 1
        line = line2 + 8+line_dict_person[n][1]+1 
        ws.cell(row = line2, column = 1).value = "2주차"
        for i in range(line2, line , 1):
            for j in range(1, 9 ,1):
                ws.cell(row = i, column = j).fill = PatternFill(fgColor = "FFFFCD", fill_type = "solid")

    if line_dict_person[n][2] != 0:        
        line2 = line + 1
        line = line2 + 8+line_dict_person[n][2]+1
        ws.cell(row = line2, column = 1).value = "3주차"
        for i in range(line2, line, 1):
            for j in range(1, 9 ,1):
                ws.cell(row = i, column = j).fill = PatternFill(fgColor = "FFEAEA", fill_type = "solid")
    if line_dict_person[n][3] != 0:        
        line2 = line + 1
        line = line2 + 8 + line_dict_person[n][3]+1    
        ws.cell(row = line2, column = 1).value = "4주차"
        for i in range(line2, line, 1):
            for j in range(1, 9 ,1):
                ws.cell(row = i, column = j).fill = PatternFill(fgColor = "F8FFFF", fill_type = "solid")
    if line_dict_person[n][4] != 0: 
        line2 = line + 1
        line = line2 + 8 + line_dict_person[n][4]+1        
        ws.cell(row = line2, column = 1).value = "5주차"
        for i in range(line2, line, 1):
            for j in range(1, 9 ,1): 
                ws.cell(row = i, column = j).fill = PatternFill(fgColor = "D9E5FF", fill_type = "solid")
            
    wb.save("./"+ str(originMonth) +"월/" + n + "_" + str(originMonth) + "월 급여내역서.xlsx")
print("#######개인별 급여 내역서 생성 완료#######")